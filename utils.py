import os
import random
from datetime import date

import openpyxl
import openpyxl.cell
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

from utils.env import env

if "nai-diffusion-4" not in env.model:
    from utils.jsondata import json_for_t2i
else:
    from utils.jsondata import json_for_t2i_v4 as json_for_t2i
from utils.utils import (
    file_path2name,
    generate_image,
    generate_random_str,
    logger,
    return_x64,
    save_image,
    sleep_for_cool,
)

README = """## 使用说明

使用前, 请打开插件目录将 "脚本文件示例.xlsx" 文件复制一份到任意位置.

然后打开复制的文件, 将要依次生成的 tag 填入 TAG 列, 保存并关闭该文件.

推文列可以不填, 为以后对接 GPT 做准备用的(实现通过推文生成 TAG 再生成图片).

回到 webui 插件页面, 设置每条 tag 生成的数量, 以及生图参数, 点击开始生成即可.

"""


def number_to_letters(n):
    result = ""
    while n >= 0:
        result = chr(n % 26 + ord("A")) + result
        n = n // 26 - 1
    return result


if not os.path.exists("./plugins/t2i/sanp_plugin_stories_to_images/脚本文件示例.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.row_dimensions[1].height = 50
    for row in range(2, 999):
        ws.row_dimensions[row].height = 300
    for col in [number_to_letters(num) for num in range(1000)]:
        ws.column_dimensions[col].width = 40
    ws.append(
        [
            "推文",
            "TAG",
            "图片",
        ]
    )
    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment
    wb.save("./plugins/t2i/sanp_plugin_stories_to_images/脚本文件示例.xlsx")


def generate(
    excel_path,
    images_number,
    sti_negative,
    sti_width,
    sti_height,
    sti_scale,
    sti_rescale,
    sti_steps,
    sti_sampler,
    sti_noise_schedule,
    sti_sm,
    sti_sm_dyn,
    sti_variety,
    sti_decrisp,
):
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook["Sheet"]
    col_num = 2
    positive = sheet[f"B{col_num}"].value
    while positive is not None:
        row_num = ord("C") - 65
        for _ in range(images_number):
            saved_path = "寄"
            while saved_path == "寄":
                json_for_t2i["input"] = positive
                json_for_t2i["parameters"]["width"] = return_x64(int(sti_width))
                json_for_t2i["parameters"]["height"] = return_x64(int(sti_height))
                json_for_t2i["parameters"]["scale"] = sti_scale
                json_for_t2i["parameters"]["cfg_rescale"] = sti_rescale
                json_for_t2i["parameters"]["sampler"] = sti_sampler
                json_for_t2i["parameters"]["steps"] = sti_steps
                if "nai-diffusion-4" not in env.model:
                    json_for_t2i["parameters"]["sm"] = (
                        sti_sm if sti_sampler != "ddim_v3" else False
                    )
                    json_for_t2i["parameters"]["sm_dyn"] = (
                        sti_sm_dyn if sti_sm and sti_sampler != "ddim_v3" else False
                    )
                json_for_t2i["parameters"]["skip_cfg_above_sigma"] = (
                    19.343056794463642
                    if "nai-diffusion-4" in env.model
                    and "nai-diffusion-4-5" not in env.model
                    else (
                        19
                        if "nai-diffusion-4-5" not in env.model
                        else 58 if sti_variety else None
                    )
                )
                json_for_t2i["parameters"]["dynamic_thresholding"] = sti_decrisp
                if sti_sampler != "ddim_v3":
                    json_for_t2i["parameters"]["noise_schedule"] = sti_noise_schedule
                seed = random.randint(1000000000, 9999999999)
                json_for_t2i["parameters"]["seed"] = seed
                json_for_t2i["parameters"]["negative_prompt"] = sti_negative

                if "nai-diffusion-4" in env.model:
                    json_for_t2i["parameters"]["use_coords"] = False
                    json_for_t2i["parameters"]["v4_prompt"]["caption"][
                        "base_caption"
                    ] = ""
                    json_for_t2i["parameters"]["v4_prompt"]["use_coords"] = False
                    json_for_t2i["parameters"]["v4_negative_prompt"]["caption"][
                        "base_caption"
                    ] = ""

                logger.debug(json_for_t2i)

                if env.save_path == "默认(Default)":
                    _path = ""
                elif env.save_path == "日期(Date)":
                    _path = f"/{date.today()}"

                if not os.path.exists(
                    path := "./output/t2i{}/{}/{}".format(
                        _path,
                        file_path2name(excel_path).split(".")[0],
                        f"B{col_num}",
                    )
                ):
                    os.makedirs(path)

                saved_path = save_image(
                    generate_image(json_for_t2i),
                    "t2i",
                    seed,
                    None,
                    None,
                    "{}/{}/{}".format(
                        file_path2name(excel_path).split(".")[0],
                        f"B{col_num}",
                        f"{seed}{generate_random_str(6)}_None_None.png",
                    ),
                )

                if saved_path == "寄":
                    sleep_for_cool(2, 4)

            image = Image(saved_path)
            w = image.width
            h = image.height
            image.width, image.height = 260, int(260 / w * h)

            sheet.add_image(image, "{}{}".format(number_to_letters(row_num), col_num))

            sleep_for_cool(env.t2i_cool_time - 3, env.t2i_cool_time + 3)

            row_num += 1
        col_num += 1
        positive = sheet[f"B{col_num}"].value

    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

    workbook.save(excel_path)
    logger.success("处理完成!")
    return "处理完成!"
